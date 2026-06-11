"""
Fast policy review workbook generator for the Caden tab.

This script keeps the original workbook untouched. It checks rows 9-315 and
columns E-BH, proposes policy value/year/link changes in a new workbook,
bolds changed cells, highlights revisions/found policies green, highlights
expired links red, and adds a Summary of Changes column.

Performance notes:
- Row-level web work runs in a ThreadPoolExecutor.
- Worker threads do not touch openpyxl objects.
- The cache is shared through a lock and saved in batches.
- The workbook is edited only on the main thread and saved atomically in batches.
"""

from __future__ import annotations

import argparse
import json
import re
import threading
import time
import urllib.parse
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from html import unescape
from pathlib import Path
from typing import Iterable

import requests

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl\n"
        "Install it with: python -m pip install openpyxl"
    ) from exc


DEFAULT_INPUT = Path(
    r"C:\Users\caden\Downloads\Summer 2026 Board Policy Indicator Refresh Data Tracker.xlsx"
)
DEFAULT_OUTPUT = Path("caden_policy_review_rows_9_315_fast.xlsx")
DEFAULT_CACHE = Path("policy_review_cache.json")
DEFAULT_PROGRESS = Path("policy_review_fast_progress.json")

SHEET_NAME = "Caden"
HEADER_ROW = 2
START_ROW = 9
END_ROW = 315
START_COL = 5  # E
END_COL = 60  # BH
GROUP_WIDTH = 4

TIMEOUT_SECONDS = 5
REQUEST_DELAY_SECONDS = 1.5
SEARCH_MAX_RESULTS = 4
SAFE_ROUTES_SEARCH_MAX_RESULTS = 10
DATABASE_SEARCH_MAX_RESULTS = 3
DEFAULT_WORKERS = 3
DEFAULT_SEARCH_INTERVAL = 2.0  # minimum seconds between any two DDG search requests
DEFAULT_CACHE_SAVE_INTERVAL = 10
DEFAULT_WORKBOOK_SAVE_INTERVAL = 10
DEFAULT_MAX_REQUESTS_PER_ROW = 200
DEFAULT_ROW_HEIGHT = 15

GREEN_FILL_NAME = "green"
RED_FILL_NAME = "red"
YELLOW_FILL_NAME = "yellow"

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
CHANGED_FONT = Font(bold=True)
LINK_ALIGNMENT = Alignment(horizontal="left", shrink_to_fit=False, wrap_text=True)

SEARCH_URL = "https://duckduckgo.com/html/?q={query}"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125 Safari/537.36"
)
POLICY_HOST_MARKERS = [
    "simbli.eboardsolutions.com",
    "go.boarddocs.com",
    "gamutonline.net",
    "boardpolicyonline.com",
    "agendaonline.net",
]
SAFE_ROUTES_CODE = "5142.2"
SAFE_ROUTES_PHRASE = "safe routes to school"

thread_local = threading.local()

# Global rate limiter: ensures all threads together fire at most 1 DDG search
# every DEFAULT_SEARCH_INTERVAL seconds, preventing anti-bot throttling.
class SearchRateLimiter:
    """Thread-safe rate limiter for DuckDuckGo HTML search requests.

    All worker threads share a single instance. A global lock ensures requests
    are serialized so DDG never sees a burst of concurrent queries.
    """

    def __init__(self, min_interval: float = DEFAULT_SEARCH_INTERVAL):
        self._lock = threading.Lock()
        self._min_interval = min_interval
        self._last_request_time: float = 0.0

    def wait(self) -> None:
        """Block until it is safe to fire the next search request."""
        with self._lock:
            now = time.monotonic()
            gap = self._min_interval - (now - self._last_request_time)
            if gap > 0:
                time.sleep(gap)
            self._last_request_time = time.monotonic()


_search_rate_limiter = SearchRateLimiter()


@dataclass(frozen=True)
class PolicyGroup:
    value_col: int
    adopted_col: int
    revised_col: int
    link_col: int
    label: str
    code: str
    title: str


@dataclass
class FetchResult:
    ok: bool
    final_url: str
    status: int | None
    text: str
    error: str | None = None


@dataclass(frozen=True)
class RowSnapshot:
    row: int
    district: str
    values: dict[int, str]


@dataclass(frozen=True)
class CellUpdate:
    col: int
    value: str
    fill: str


@dataclass(frozen=True)
class CellMark:
    col: int
    fill: str


@dataclass
class RowResult:
    row: int
    notes: list[str] = field(default_factory=list)
    updates: list[CellUpdate] = field(default_factory=list)
    marks: list[CellMark] = field(default_factory=list)


@dataclass
class RequestBudget:
    remaining: int

    def take(self) -> bool:
        if self.remaining <= 0:
            return False
        self.remaining -= 1
        return True


class ThreadSafeCache:
    def __init__(self, path: Path):
        self.path = path
        self.lock = threading.Lock()
        self.dirty = False
        self.data = self._load()

    def _load(self) -> dict:
        if not self.path.exists():
            return {}
        return json.loads(self.path.read_text(encoding="utf-8"))

    def get(self, key: str):
        with self.lock:
            return self.data.get(key)

    def set(self, key: str, value) -> None:
        with self.lock:
            self.data[key] = value
            self.dirty = True

    def save_if_dirty(self) -> None:
        with self.lock:
            if not self.dirty:
                return
            payload = json.dumps(self.data, indent=2, sort_keys=True)
            self.dirty = False
        tmp_path = self.path.with_name(f"{self.path.stem}.tmp{self.path.suffix}")
        tmp_path.write_text(payload, encoding="utf-8")
        tmp_path.replace(self.path)


def get_session() -> requests.Session:
    session = getattr(thread_local, "session", None)
    if session is None:
        session = requests.Session()
        session.headers.update({"User-Agent": USER_AGENT})
        thread_local.session = session
    return session


def normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def is_url(value: str) -> bool:
    return value.lower().startswith(("http://", "https://"))


def extract_policy_parts(label: str) -> tuple[str, str]:
    match = re.match(r"^(BP|AR)\s+([0-9.]+)(?:\s+[-\u2013]\s+|\s+)?(.*)$", label, re.IGNORECASE)
    if match:
        return f"{match.group(1).upper()} {match.group(2)}", match.group(3).strip()
    return "", label.strip()


def is_safe_routes_policy(policy: PolicyGroup) -> bool:
    return SAFE_ROUTES_CODE in policy.code or SAFE_ROUTES_PHRASE in policy.title.lower()


def policy_code_variants(policy: PolicyGroup) -> list[str]:
    variants = [policy.code.lower()]
    if is_safe_routes_policy(policy):
        doc_type = policy.code.split()[0].lower() if policy.code else ""
        variants.extend(
            [
                SAFE_ROUTES_CODE,
                f"{SAFE_ROUTES_CODE}(a)",
                f"{SAFE_ROUTES_CODE} (a)",
                f"{doc_type} {SAFE_ROUTES_CODE}",
                f"{doc_type} {SAFE_ROUTES_CODE}(a)",
                f"{doc_type} {SAFE_ROUTES_CODE} (a)",
            ]
        )
    deduped: list[str] = []
    for variant in variants:
        if variant and variant not in deduped:
            deduped.append(variant)
    return deduped


def normalized_policy_code_variants(policy: PolicyGroup) -> list[str]:
    return [
        re.sub(r"[\s._%()/-]+", "", variant)
        for variant in policy_code_variants(policy)
    ]


def request_url(url: str, timeout: int, budget: RequestBudget | None = None) -> FetchResult:
    if budget is not None and not budget.take():
        return FetchResult(False, url, None, "", "row request budget exhausted")
    try:
        response = get_session().get(url, timeout=(3, timeout), allow_redirects=True)
        response.encoding = response.encoding or "utf-8"
        return FetchResult(
            ok=200 <= response.status_code < 400,
            final_url=response.url,
            status=response.status_code,
            text=response.text[:800_000],
        )
    except Exception as exc:
        return FetchResult(False, url, None, "", str(exc))


def cached_fetch(
    url: str,
    cache: ThreadSafeCache,
    timeout: int,
    delay: float,
    budget: RequestBudget | None = None,
) -> FetchResult:
    key = f"fetch::{url}"
    cached = cache.get(key)
    if cached is not None:
        return FetchResult(**cached)
    result = request_url(url, timeout, budget)
    is_bot_blocked = "Pardon Our Interruption" in result.text and "bot" in result.text.lower()
    if not is_bot_blocked:
        cache.set(key, result.__dict__)
    if delay:
        time.sleep(delay)
    return result


def page_looks_expired(result: FetchResult) -> bool:
    if result.status in (404, 410):
        return True
    if not result.ok and result.status not in (401, 403):
        return True

    text = result.text.lower()
    expired_markers = [
        "404 not found",
        "page cannot be found",
        "policy not found",
        "policy is no longer available",
        "error 404",
    ]
    return any(marker in text for marker in expired_markers)


def visible_text(html: str) -> str:
    html = re.sub(r"(?is)<script.*?</script>|<style.*?</style>", " ", html)
    html = re.sub(r"(?s)<[^>]+>", " ", html)
    return re.sub(r"\s+", " ", unescape(html)).strip()


def extract_years(text: str) -> tuple[str | None, str | None]:
    adopted = None
    revised = None
    # Patterns handle both bare-year format ("Adopted: 2012") and
    # MM/DD/YYYY format as used by Simbli ("Original Adopted Date: 09/06/2012").
    # The optional (?:Date)? absorbs the word "Date" after the label.
    # The optional (?:\d{1,2}[/.-]\d{1,2}[/.-])? absorbs "MM/DD/" or "MM-DD-".
    adopted_patterns = [
        r"\bAdopted\s*(?:Date)?\s*:?\s*(?:[^0-9]{0,20})?(?:\d{1,2}[/.-]\d{1,2}[/.-])?((?:19|20)\d{2})",
        r"\bApproved\s*(?:Date)?\s*:?\s*(?:[^0-9]{0,20})?(?:\d{1,2}[/.-]\d{1,2}[/.-])?((?:19|20)\d{2})",
        r"\bOriginal\s+Adopted\s*(?:Date)?\s*:?\s*(?:\d{1,2}[/.-]\d{1,2}[/.-])?((?:19|20)\d{2})",
    ]
    revised_patterns = [
        r"\bLast\s+Revised\s*(?:Date)?\s*:?\s*(?:\d{1,2}[/.-]\d{1,2}[/.-])?((?:19|20)\d{2})",
        r"\bRevised\s*(?:Date)?\s*:?\s*(?:[^0-9]{0,20})?(?:\d{1,2}[/.-]\d{1,2}[/.-])?((?:19|20)\d{2})",
        r"\bUpdated\s*(?:Date)?\s*:?\s*(?:[^0-9]{0,20})?(?:\d{1,2}[/.-]\d{1,2}[/.-])?((?:19|20)\d{2})",
        r"\bLast\s+Reviewed\s*(?:Date)?\s*:?\s*(?:\d{1,2}[/.-]\d{1,2}[/.-])?((?:19|20)\d{2})",
    ]
    for pattern in adopted_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            adopted = match.group(1)
            break
    revised_matches: list[str] = []
    for pattern in revised_patterns:
        revised_matches.extend(re.findall(pattern, text, re.IGNORECASE))
    if revised_matches:
        revised = max(revised_matches)
    return adopted, revised



def html_search(
    query: str,
    cache: ThreadSafeCache,
    timeout: int,
    delay: float,
    budget: RequestBudget | None = None,
    max_results: int = SEARCH_MAX_RESULTS,
) -> list[str]:
    key = f"search::{query}"
    cached = cache.get(key)
    if cached is not None:
        return cached

    # Gate all live DDG requests through the global rate limiter so threads
    # cannot pile up simultaneous requests and trigger anti-bot blocking.
    _search_rate_limiter.wait()

    url = SEARCH_URL.format(query=urllib.parse.quote_plus(query))
    result = request_url(url, timeout, budget)
    links: list[str] = []
    if result.ok:
        for raw in re.findall(r'href="([^"]+)"', result.text):
            raw = unescape(raw)
            parsed = urllib.parse.urlparse(raw)
            if parsed.path == "/l/":
                params = urllib.parse.parse_qs(parsed.query)
                raw = params.get("uddg", [raw])[0]
            if raw.startswith(("http://", "https://")) and "duckduckgo.com" not in raw:
                links.append(raw)

    deduped = []
    seen = set()
    for link in links:
        clean = link.split("&rut=")[0]
        if clean not in seen:
            deduped.append(clean)
            seen.add(clean)
        if len(deduped) >= max_results:
            break

    cache.set(key, deduped)
    if delay:
        time.sleep(delay)
    return deduped


def search_queries(district: str, policy: PolicyGroup) -> Iterable[str]:
    quoted_district = f'"{district}"'
    if is_safe_routes_policy(policy):
        yield f'{quoted_district} "{SAFE_ROUTES_PHRASE}" {SAFE_ROUTES_CODE}'
        yield f'{quoted_district} {SAFE_ROUTES_PHRASE} {SAFE_ROUTES_CODE}'
        yield f'{quoted_district} "{SAFE_ROUTES_PHRASE}" "board policy"'
        yield f'{quoted_district} {policy.code}'
        yield f'{quoted_district} {policy.code} {SAFE_ROUTES_PHRASE} policy'
        yield f'{quoted_district} {policy.code} site:simbli.eboardsolutions.com'
        yield f'{quoted_district} {policy.code} site:go.boarddocs.com'
        yield f'{quoted_district} {policy.code} site:gamutonline.net'
        yield f'{quoted_district} "{SAFE_ROUTES_PHRASE}" site:simbli.eboardsolutions.com'
        yield f'{quoted_district} "{SAFE_ROUTES_PHRASE}" site:go.boarddocs.com'
        yield f'{quoted_district} "{SAFE_ROUTES_PHRASE}" site:gamutonline.net'
        return

    yield f'{quoted_district} "{policy.code}" "{policy.title}" board policy'
    yield f'{quoted_district} "{policy.code}" site:simbli.eboardsolutions.com'
    yield f'{quoted_district} "{policy.code}" site:go.boarddocs.com'


def result_matches_policy(url: str, text: str, policy: PolicyGroup) -> bool:
    haystack = f"{url} {text}".lower()
    compact = re.sub(r"[\s._%()/-]+", "", haystack)
    code_match = any(
        variant in haystack or normalized in compact
        for variant, normalized in zip(policy_code_variants(policy), normalized_policy_code_variants(policy))
    )
    if is_safe_routes_policy(policy):
        phrase_match = (
            SAFE_ROUTES_PHRASE in haystack
            or "safe routes" in haystack
            or "saferoutes" in compact
        )
        policy_context = any(
            term in compact
            for term in [
                "boardpolicy",
                "administrativeregulation",
                "policy51422",
                "regulation51422",
                "bp51422",
                "ar51422",
                "51422a",
                "51422",
                "policy",
                "regulation"
            ]
        )
        return code_match or (phrase_match and policy_context)
    return code_match


def candidate_is_policy_system(url: str, policy: PolicyGroup) -> bool:
    lowered = url.lower()
    compact_url = lowered.replace("%20", "").replace("-", "").replace("_", "")
    code_match = any(normalized in compact_url for normalized in normalized_policy_code_variants(policy))
    safe_routes_url_match = is_safe_routes_policy(policy) and "saferoutestoschool" in compact_url
    if is_safe_routes_policy(policy):
        return True
    return any(marker in lowered for marker in POLICY_HOST_MARKERS) or code_match or safe_routes_url_match


def find_policy_link(
    district: str,
    policy: PolicyGroup,
    cache: ThreadSafeCache,
    timeout: int,
    delay: float,
    budget: RequestBudget,
) -> tuple[str | None, str | None, str | None]:
    for query in search_queries(district, policy):
        max_results = SAFE_ROUTES_SEARCH_MAX_RESULTS if is_safe_routes_policy(policy) else SEARCH_MAX_RESULTS
        for candidate in html_search(query, cache, timeout, delay, budget, max_results):
            if not candidate_is_policy_system(candidate, policy):
                continue
            fetched = cached_fetch(candidate, cache, timeout, delay, budget)
            if page_looks_expired(fetched):
                continue
            text = visible_text(fetched.text)
            if result_matches_policy(fetched.final_url, text, policy):
                adopted, revised = extract_years(text)
                return fetched.final_url, adopted, revised
    return None, None, None


def district_policy_database_exists(
    district: str,
    cache: ThreadSafeCache,
    timeout: int,
    delay: float,
    budget: RequestBudget,
) -> bool:
    queries = [
        f'"{district}" "board policy"',
        f'"{district}" "simbli"',
        f'"{district}" "BoardDocs"',
    ]
    for query in queries:
        for link in html_search(query, cache, timeout, delay, budget, DATABASE_SEARCH_MAX_RESULTS):
            lowered = link.lower()
            if any(host in lowered for host in POLICY_HOST_MARKERS):
                return True
    return False


def as_int_year(value: str) -> int | None:
    match = re.search(r"(19|20)\d{2}", value)
    if not match:
        return None
    return int(match.group(0))


def add_update(result: RowResult, snapshot: RowSnapshot, col: int, value: str, fill: str) -> bool:
    old = snapshot.values.get(col, "")
    if old == normalize_cell(value):
        return False
    result.updates.append(CellUpdate(col, value, fill))
    return True


def add_mark(result: RowResult, col: int, fill: str) -> None:
    result.marks.append(CellMark(col, fill))


def process_existing_policy(
    result: RowResult,
    snapshot: RowSnapshot,
    policy: PolicyGroup,
    cache: ThreadSafeCache,
    timeout: int,
    delay: float,
    budget: RequestBudget,
) -> None:
    district = snapshot.district
    link = snapshot.values.get(policy.link_col, "")
    old_adopted = snapshot.values.get(policy.adopted_col, "")
    old_revised = snapshot.values.get(policy.revised_col, "")

    if not is_url(link):
        found_link, found_adopted, found_revised = find_policy_link(
            district, policy, cache, timeout, delay, budget
        )
        if found_link:
            add_update(result, snapshot, policy.link_col, found_link, GREEN_FILL_NAME)
            if found_adopted:
                add_update(result, snapshot, policy.adopted_col, found_adopted, GREEN_FILL_NAME)
            if found_revised:
                add_update(result, snapshot, policy.revised_col, found_revised, GREEN_FILL_NAME)
            result.notes.append(f"{policy.code}: found replacement link for existing policy")
        return

    fetched = cached_fetch(link, cache, timeout, delay, budget)
    if page_looks_expired(fetched):
        add_mark(result, policy.value_col, RED_FILL_NAME)
        add_mark(result, policy.link_col, RED_FILL_NAME)
        found_link, found_adopted, found_revised = find_policy_link(
            district, policy, cache, timeout, delay, budget
        )
        if found_link:
            add_update(result, snapshot, policy.link_col, found_link, RED_FILL_NAME)
            if found_adopted:
                add_update(result, snapshot, policy.adopted_col, found_adopted, RED_FILL_NAME)
            if found_revised:
                add_update(result, snapshot, policy.revised_col, found_revised, RED_FILL_NAME)
            result.notes.append(f"{policy.code}: original link expired; replacement found")
        else:
            replacement = "N/A" if district_policy_database_exists(
                district, cache, timeout, delay, budget
            ) else "*"
            add_update(result, snapshot, policy.link_col, replacement, RED_FILL_NAME)
            result.notes.append(f"{policy.code}: original link expired; no replacement found")
        return

    text = visible_text(fetched.text)
    adopted, revised = extract_years(text)
    changed_parts: list[str] = []
    if adopted and old_adopted in {"", "N/A", "*"}:
        if add_update(result, snapshot, policy.adopted_col, adopted, GREEN_FILL_NAME):
            changed_parts.append(f"adopted year {old_adopted or 'blank'} -> {adopted}")
    if revised:
        old_year = as_int_year(old_revised)
        new_year = as_int_year(revised)
        if new_year and (old_year is None or new_year > old_year):
            if add_update(result, snapshot, policy.revised_col, revised, GREEN_FILL_NAME):
                changed_parts.append(f"revised year {old_revised or 'blank'} -> {revised}")
    if fetched.final_url and fetched.final_url != link:
        if add_update(result, snapshot, policy.link_col, fetched.final_url, GREEN_FILL_NAME):
            changed_parts.append("canonical link updated")

    if changed_parts:
        add_mark(result, policy.value_col, GREEN_FILL_NAME)
        result.notes.append(f"{policy.code}: " + "; ".join(changed_parts))


def process_missing_policy(
    result: RowResult,
    snapshot: RowSnapshot,
    policy: PolicyGroup,
    cache: ThreadSafeCache,
    timeout: int,
    delay: float,
    budget: RequestBudget,
) -> None:
    found_link, adopted, revised = find_policy_link(
        snapshot.district, policy, cache, timeout, delay, budget
    )
    if not found_link:
        return
    add_update(result, snapshot, policy.value_col, "1", GREEN_FILL_NAME)
    add_update(result, snapshot, policy.link_col, found_link, GREEN_FILL_NAME)
    if adopted:
        add_update(result, snapshot, policy.adopted_col, adopted, GREEN_FILL_NAME)
    if revised:
        add_update(result, snapshot, policy.revised_col, revised, GREEN_FILL_NAME)
    result.notes.append(f"{policy.code}: policy found and proposed as passed")


def process_row(
    snapshot: RowSnapshot,
    groups: list[PolicyGroup],
    cache: ThreadSafeCache,
    timeout: int,
    delay: float,
    max_requests_per_row: int,
) -> RowResult:
    result = RowResult(row=snapshot.row)
    budget = RequestBudget(max_requests_per_row)
    # Process Safe Routes groups first so they are never budget-starved by earlier columns
    prioritized = sorted(groups, key=lambda g: (0 if is_safe_routes_policy(g) else 1))
    for policy in prioritized:
        value = snapshot.values.get(policy.value_col, "")
        if value in {"1", "1.0"}:
            process_existing_policy(result, snapshot, policy, cache, timeout, delay, budget)
        else:
            process_missing_policy(result, snapshot, policy, cache, timeout, delay, budget)
        if budget.remaining <= 0:
            result.notes.append("Row request budget exhausted before all policy checks completed")
            break
    return result


def build_policy_groups(ws) -> list[PolicyGroup]:
    groups: list[PolicyGroup] = []
    for col in range(START_COL, END_COL + 1, GROUP_WIDTH):
        label = normalize_cell(ws.cell(HEADER_ROW, col).value)
        code, title = extract_policy_parts(label)
        groups.append(
            PolicyGroup(
                value_col=col,
                adopted_col=col + 1,
                revised_col=col + 2,
                link_col=col + 3,
                label=label,
                code=code,
                title=title,
            )
        )
    return groups


def build_snapshots(ws, start_row: int, end_row: int) -> list[RowSnapshot]:
    snapshots: list[RowSnapshot] = []
    for row in range(start_row, end_row + 1):
        district = normalize_cell(ws.cell(row, 4).value)
        if not district:
            continue
        values = {col: normalize_cell(ws.cell(row, col).value) for col in range(START_COL, END_COL + 1)}
        snapshots.append(RowSnapshot(row=row, district=district, values=values))
    return snapshots


def policy_link_columns() -> list[int]:
    return list(range(START_COL + GROUP_WIDTH - 1, END_COL + 1, GROUP_WIDTH))


def apply_link_layout(ws, start_row: int, end_row: int) -> None:
    for col in policy_link_columns():
        ws.column_dimensions[get_column_letter(col)].width = 24
        for row in range(HEADER_ROW, end_row + 1):
            ws.cell(row, col).alignment = LINK_ALIGNMENT

    for row in range(start_row, end_row + 1):
        ws.row_dimensions[row].height = DEFAULT_ROW_HEIGHT


def fill_for_name(name: str) -> PatternFill:
    if name == GREEN_FILL_NAME:
        return GREEN_FILL
    if name == RED_FILL_NAME:
        return RED_FILL
    if name == YELLOW_FILL_NAME:
        return YELLOW_FILL
    raise ValueError(f"Unknown fill name: {name}")


def mark_cell(cell, fill_name: str) -> None:
    cell.font = CHANGED_FONT
    cell.fill = fill_for_name(fill_name)


def apply_row_result(ws, result: RowResult, summary_col: int) -> None:
    for mark in result.marks:
        mark_cell(ws.cell(result.row, mark.col), mark.fill)
    for update in result.updates:
        cell = ws.cell(result.row, update.col)
        cell.value = update.value
        mark_cell(cell, update.fill)
    if result.notes:
        summary_cell = ws.cell(result.row, summary_col)
        summary_cell.value = " | ".join(result.notes)
        mark_cell(summary_cell, YELLOW_FILL_NAME)


def save_workbook_atomic(wb, output_path: Path) -> None:
    tmp_path = output_path.with_name(f"{output_path.stem}.tmp{output_path.suffix}")
    wb.save(tmp_path)
    tmp_path.replace(output_path)


def save_progress_atomic(path: Path, completed_rows: set[int], total_rows: int) -> None:
    payload = {
        "completed_count": len(completed_rows),
        "total_count": total_rows,
        "completed_rows": sorted(completed_rows),
        "last_updated": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    tmp_path = path.with_name(f"{path.stem}.tmp{path.suffix}")
    tmp_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    tmp_path.replace(path)


def generate_review(
    input_path: Path,
    output_path: Path,
    cache_path: Path,
    start_row: int,
    end_row: int,
    workers: int,
    cache_save_interval: int,
    workbook_save_interval: int,
    timeout: int,
    delay: float,
    progress_path: Path,
    max_requests_per_row: int,
) -> None:
    wb = load_workbook(input_path)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {SHEET_NAME}")
    ws = wb[SHEET_NAME]
    snapshots = build_snapshots(ws, start_row, end_row)
    groups = build_policy_groups(ws)
    cache = ThreadSafeCache(cache_path)

    summary_col = END_COL + 1
    ws.insert_cols(summary_col)
    ws.cell(HEADER_ROW, summary_col).value = "Summary of Changes"
    ws.cell(HEADER_ROW, summary_col).font = Font(bold=True)
    ws.column_dimensions[get_column_letter(summary_col)].width = 80
    apply_link_layout(ws, start_row, end_row)

    completed = 0
    completed_rows: set[int] = set()
    print(f"Queued {len(snapshots)} district rows with {workers} workers.", flush=True)
    try:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_to_snapshot = {
                executor.submit(
                    process_row,
                    snapshot,
                    groups,
                    cache,
                    timeout,
                    delay,
                    max_requests_per_row,
                ): snapshot
                for snapshot in snapshots
            }
            for future in as_completed(future_to_snapshot):
                snapshot = future_to_snapshot[future]
                try:
                    result = future.result()
                except Exception as exc:
                    print(f"Row {snapshot.row}: {snapshot.district} failed: {exc}", flush=True)
                    continue

                apply_row_result(ws, result, summary_col)
                completed += 1
                completed_rows.add(snapshot.row)
                print(
                    f"Completed {completed}/{len(snapshots)}: "
                    f"row {snapshot.row} {snapshot.district} "
                    f"({len(result.notes)} note(s))",
                    flush=True,
                )

                if completed % cache_save_interval == 0:
                    cache.save_if_dirty()
                    save_progress_atomic(progress_path, completed_rows, len(snapshots))
                if completed % workbook_save_interval == 0:
                    save_workbook_atomic(wb, output_path)
    finally:
        cache.save_if_dirty()
        save_progress_atomic(progress_path, completed_rows, len(snapshots))
        save_workbook_atomic(wb, output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate Caden policy review workbook quickly.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--cache", type=Path, default=DEFAULT_CACHE)
    parser.add_argument("--progress", type=Path, default=DEFAULT_PROGRESS)
    parser.add_argument("--start-row", type=int, default=START_ROW)
    parser.add_argument("--end-row", type=int, default=END_ROW)
    parser.add_argument("--workers", type=int, default=DEFAULT_WORKERS)
    parser.add_argument("--cache-save-interval", type=int, default=DEFAULT_CACHE_SAVE_INTERVAL)
    parser.add_argument("--workbook-save-interval", type=int, default=DEFAULT_WORKBOOK_SAVE_INTERVAL)
    parser.add_argument("--timeout", type=int, default=TIMEOUT_SECONDS)
    parser.add_argument("--delay", type=float, default=REQUEST_DELAY_SECONDS)
    parser.add_argument("--max-requests-per-row", type=int, default=DEFAULT_MAX_REQUESTS_PER_ROW)
    args = parser.parse_args()

    generate_review(
        input_path=args.input,
        output_path=args.output,
        cache_path=args.cache,
        start_row=args.start_row,
        end_row=args.end_row,
        workers=args.workers,
        cache_save_interval=args.cache_save_interval,
        workbook_save_interval=args.workbook_save_interval,
        timeout=args.timeout,
        delay=args.delay,
        progress_path=args.progress,
        max_requests_per_row=args.max_requests_per_row,
    )
    print(f"Review workbook written to: {args.output.resolve()}", flush=True)


if __name__ == "__main__":
    main()

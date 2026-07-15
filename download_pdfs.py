"""
download_pdfs.py — Download a PDF for every adopted policy in a scraped workbook.

Input  : a .xlsx workbook produced by the scraper (e.g. Summer_2026_Scraped_*.xlsx)
Output : a ``Ten Strands`` parent folder in the current directory, containing a
         ``By Policy`` subfolder with one folder per policy (e.g.
         ``BP 3510 Green Schools Operations``), containing every district's
         downloaded PDF for that policy. Policy folder names are taken from a
         canonical list (CANONICAL_POLICIES below) keyed on policy type +
         number, so title-wording differences across districts don't create
         duplicate folders. Any policy code not in that list still gets a
         folder, auto-named from its own code and title.

Naming convention (one file per adopted policy):
    <BP/AR>_<Policy Number>_<Policy Title>_<County>_<District>_<Year Adopted>_<Year Revised>.pdf
Example:
    BP_3510_Green Schools Operations_Alameda_Alameda Unified_2019_2019.pdf

Only BP/AR policy quads are considered. Resolution columns (RES-*) are skipped.

How PDFs are produced
---------------------
The workbook link cell is usually an HTML page, not a PDF:
  - Simbli   : .../Policy/ViewPolicy.aspx?S=...   (HTML, behind Cloudflare)
  - BoardDocs: .../Board.nsf/goto?open&id=...      (HTML)
  - District : https://berkeleyschools.net/...pdf   (already a PDF)

This script uses a hybrid, Chrome-centric strategy:
  - If the URL ends in ``.pdf``  -> download the original file directly (lossless).
  - Otherwise                    -> load the page in undetected-chromedriver
                                     (bypasses Cloudflare, same engine the scraper
                                     uses) and print it to PDF via Chrome DevTools
                                     ``Page.printToPDF``. This works uniformly for
                                     Simbli, BoardDocs, and any other platform.

Usage
-----
    python download_pdfs.py --input "Summer_2026_Scraped_20260702_095536.xlsx" --sheet Caden
    python download_pdfs.py --input ... --sheet Caden --limit 5                 # test run
    python download_pdfs.py --input ... --sheet Caden --start-row 3 --end-row 20
    python download_pdfs.py --input ... --sheet Caden --overwrite              # re-download
    python download_pdfs.py --input ... --sheet Caden --dry-run                # preview only
"""

from __future__ import annotations

import argparse
import asyncio
import base64
import json
import os
import random
import re
import sys
import time
from datetime import datetime
from typing import Optional

import httpx
import openpyxl

# Column layout (value | adopted | revised | link quads) is shared with the
# scraper so the two never drift apart.
from models import POLICY_DEFS


# ── Constants ─────────────────────────────────────────────────────────────────

PARENT_FOLDER = "Ten Strands"
POLICY_SUBFOLDER = "By Policy"

# Canonical (doctype, policy number) -> folder name. Keeps every district's PDF
# for the same policy in one folder regardless of small title-wording
# differences between districts. Policy codes not found here still get a
# folder — it's just auto-named from the code + title instead (see
# get_policy_folder_name below).
CANONICAL_POLICIES = {
    ("AR", "3511.1"): "AR 3511.1 Integrated Waste Management",
    ("AR", "3514"): "AR 3514 Environmental Safety",
    ("AR", "3514.1"): "AR 3514.1 Hazardous Substances",
    ("AR", "3514.2"): "AR 3514.2 Integrated Pest Management",
    ("AR", "5142.2"): "AR 5142.2 Safe Routes to School",
    ("AR", "7110"): "AR 7110 Facilities Master Plan",
    ("BP", "3510"): "BP 3510 Green Schools Operations",
    ("BP", "3511"): "BP 3511 Energy and Water Management",
    ("BP", "3511.1"): "BP 3511.1 Integrated Waste Management",
    ("BP", "3514"): "BP 3514 Environmental Safety",
    ("BP", "3514.1"): "BP 3514.1 Hazardous Substances",
    ("BP", "5142.2"): "BP 5142.2 Safe Routes to School",
    ("BP", "6142.5"): "BP 6142.5 Environmental Education",
    ("BP", "7110"): "BP 7110 Facilities Master Plan",
}


# ── Filename helpers ──────────────────────────────────────────────────────────

def strip_county_suffix(county: str) -> str:
    """'Alameda County' -> 'Alameda'  (case-insensitive)."""
    return re.sub(r"\s*county\s*$", "", (county or "").strip(), flags=re.IGNORECASE)


def sanitize(value: str) -> str:
    """Remove characters that are illegal in file-system names and tidy spaces."""
    value = (value or "").strip()
    value = re.sub(r'[<>:"/\\|?*]', " ", value)
    value = re.sub(r"\s{2,}", " ", value)
    return value.strip()


def normalize_year(value) -> str:
    """'2019', 2019, 2019.0 -> '2019'; anything else -> 'NA'."""
    if value is None:
        return "NA"
    text = str(value).strip()
    if not text or text in {"N/A", "NA", "*", "None"}:
        return "NA"
    try:
        year = int(float(text))
        return str(year) if 1900 < year < 2100 else "NA"
    except (ValueError, TypeError):
        return "NA"


def parse_doc_type_and_number(policy_code: str) -> tuple[str, str]:
    """
    'BP 3510'    -> ('BP', '3510')
    'AR 3514.1'  -> ('AR', '3514.1')
    'RES-CLIMATE'-> ('', '')          (resolutions are skipped by the caller)
    """
    code = (policy_code or "").strip()
    m = re.match(r"^(BP|AR)\s+(.+)$", code, re.IGNORECASE)
    if m:
        return m.group(1).upper(), m.group(2).strip()
    return "", ""


def build_filename(policy_code: str, policy_title: str, county: str,
                   district: str, year_adopted, year_revised) -> Optional[str]:
    """Assemble the convention filename stem, or None if it cannot be built."""
    doc_type, policy_number = parse_doc_type_and_number(policy_code)
    if not doc_type or not policy_number:
        # Resolution or unknown code — caller filters these out, but be safe.
        return None
    parts = [
        doc_type,
        sanitize(policy_number),
        sanitize(policy_title) or "Untitled",
        sanitize(strip_county_suffix(county)),
        sanitize(district),
        normalize_year(year_adopted),
        normalize_year(year_revised),
    ]
    return "_".join(parts) + ".pdf"


def get_policy_folder_name(policy_code: str, policy_title: str) -> Optional[str]:
    """
    Resolve the destination policy folder name for a policy code.

    Looks up (doc_type, policy_number) in CANONICAL_POLICIES first. If not
    found, auto-generates a folder name from the code + title so the policy
    still gets its own folder instead of being skipped.
    Returns None only if the code isn't a valid BP/AR code (e.g. resolutions).
    """
    doc_type, policy_number = parse_doc_type_and_number(policy_code)
    if not doc_type or not policy_number:
        return None
    canonical = CANONICAL_POLICIES.get((doc_type, policy_number))
    if canonical:
        return canonical
    auto_name = f"{doc_type} {policy_number} {sanitize(policy_title) or 'Untitled'}"
    return sanitize(auto_name)


# ── Workbook predicates ───────────────────────────────────────────────────────

def _is_adopted(value) -> bool:
    """True if the policy value cell marks the policy as adopted (value == 1)."""
    if value is None:
        return False
    return str(value).strip() in {"1", "1.0"}


def _is_real_link(link) -> bool:
    """True if the link cell is a real http(s) URL (not N/A / * / blank)."""
    if not link:
        return False
    text = str(link).strip()
    if not text or text in {"N/A", "*", "None"}:
        return False
    return text.lower().startswith(("http://", "https://"))


# ── Workbook reading ──────────────────────────────────────────────────────────

def load_tasks(filepath: str, sheet: str,
               start_row: Optional[int] = None,
               end_row: Optional[int] = None) -> list[dict]:
    """
    Walk the sheet and return one task dict per adopted policy that has a real link.

    Each task: {row, county, district, policy_code, policy_title,
                year_adopted, year_revised, link}

    Rows 1-2 are headers; data starts at row 3 (matching the scraper's reader).
    Columns: A=status, B=CDS code, C=county, D=district, then per-policy quads
    (value | year_adopted | year_revised | link) per POLICY_DEFS.

    Only BP/AR policies are considered; resolution columns (RES-*) are skipped.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet]

    tasks: list[dict] = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=3, values_only=True), start=3
    ):
        if start_row and row_idx < start_row:
            continue
        if end_row and row_idx > end_row:
            break

        county = str(row[2]).strip() if row[2] is not None else ""
        district = str(row[3]).strip() if row[3] is not None else ""
        if not district:
            continue

        for pdef in POLICY_DEFS:
            policy_code = pdef["code"]
            # Only BP/AR policies — skip RES-* resolution columns.
            if not re.match(r"^(BP|AR)\s", policy_code):
                continue

            base = pdef["col_start"] - 1  # 0-based index into the row tuple
            value = row[base]
            year_adopted = row[base + 1]
            year_revised = row[base + 2]
            link = row[base + 3]

            if not _is_adopted(value) or not _is_real_link(link):
                continue

            tasks.append({
                "row": row_idx,
                "county": county,
                "district": district,
                "policy_code": policy_code,
                "policy_title": pdef["title"],
                "year_adopted": year_adopted,
                "year_revised": year_revised,
                "link": str(link).strip(),
            })
    return tasks


# ── Download routing ──────────────────────────────────────────────────────────

# Realistic browser headers for direct PDF downloads (some hosts block default UA).
BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "application/pdf,*/*;q=0.8"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}


def is_direct_pdf(url: str) -> bool:
    """Heuristic: a URL whose path ends in .pdf is downloaded as-is."""
    path = url.split("?", 1)[0].split("#", 1)[0]
    return path.lower().endswith(".pdf")


async def _download_direct_async(url: str, dest_path: str) -> None:
    """Stream a direct PDF link to disk with browser-like headers."""
    async with httpx.AsyncClient(
        follow_redirects=True, timeout=60.0, headers=BROWSER_HEADERS
    ) as client:
        async with client.stream("GET", url) as resp:
            resp.raise_for_status()
            with open(dest_path, "wb") as fh:
                async for chunk in resp.aiter_bytes():
                    fh.write(chunk)


# ── Chrome HTML → PDF rendering ───────────────────────────────────────────────

# Pacing between downloads — keeps requests human-paced to avoid anti-bot blocks.
_PACE_MIN = 1.5
_PACE_MAX = 3.0


def _wait_for_render(driver, url: str, timeout: float = 25.0) -> bool:
    """
    Wait for the page to clear Cloudflare/anti-bot checks and render.
    Returns True if the page looks like it loaded normally.
    """
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            title = driver.title or ""
        except Exception:
            title = ""
        try:
            src = driver.page_source or ""
            html_len = len(src)
        except Exception:
            html_len = 0
        if "Just a moment" in title or "cf-browser-verification" in src[:2000]:
            time.sleep(1.5)
            continue
        # Give Angular/SPA content a moment to populate after the shell loads.
        if html_len > 3000:
            time.sleep(2.0)
            return True
        time.sleep(1.0)
    return False


def render_pdf_via_chrome(driver, url: str, dest_path: str) -> None:
    """
    Load ``url`` in Chrome and save a PDF of the rendered page via DevTools.
    Used for Simbli, BoardDocs, and any non-.pdf HTML page.
    """
    driver.get(url)
    # Generous wait — Simbli sits behind Cloudflare and renders via Angular.
    _wait_for_render(driver, url, timeout=30.0)
    time.sleep(2.0)  # let late-rendering content settle

    result = driver.execute_cdp_cmd(
        "Page.printToPDF",
        {
            "printBackground": True,
            "preferCSSPageSize": True,
            "marginTop": 0.4,
            "marginBottom": 0.4,
            "marginLeft": 0.4,
            "marginRight": 0.4,
        },
    )
    pdf_bytes = base64.b64decode(result["data"])
    if not pdf_bytes:
        raise RuntimeError("Chrome returned an empty PDF")
    with open(dest_path, "wb") as fh:
        fh.write(pdf_bytes)


# ── Per-task processing ───────────────────────────────────────────────────────

def unique_path(dest_path: str) -> str:
    """If dest_path exists, append _1, _2, ... until free."""
    if not os.path.exists(dest_path):
        return dest_path
    stem, ext = os.path.splitext(dest_path)
    counter = 1
    while True:
        candidate = f"{stem}_{counter}{ext}"
        if not os.path.exists(candidate):
            return candidate
        counter += 1


def process_task(driver, task: dict, base_dir: str, overwrite: bool,
                 dry_run: bool = False) -> dict:
    """
    Download one policy PDF. Returns a result record for the manifest.
    status ∈ {downloaded, exists, failed}

    Direct-PDF links are streamed to disk. HTML pages (Simbli, BoardDocs,
    district sites) are rendered to PDF via Chrome's Page.printToPDF.
    If dry_run=True no file is written and status is set to 'downloaded'.
    """
    record = {
        **{k: task[k] for k in ("row", "county", "district", "policy_code",
                                 "policy_title", "link")},
        "dest": None,
        "method": None,
        "status": None,
        "error": None,
    }

    filename = build_filename(
        task["policy_code"], task["policy_title"], task["county"],
        task["district"], task["year_adopted"], task["year_revised"],
    )
    if not filename:
        record["status"] = "failed"
        record["error"] = "Could not build filename (non-BP/AR code?)"
        return record

    policy_folder_name = get_policy_folder_name(task["policy_code"], task["policy_title"])
    if not policy_folder_name:
        record["status"] = "failed"
        record["error"] = "Could not resolve policy folder (non-BP/AR code?)"
        return record

    policy_folder = os.path.join(base_dir, POLICY_SUBFOLDER, policy_folder_name)
    dest_path = os.path.join(policy_folder, filename)
    record["dest"] = dest_path

    if os.path.exists(dest_path) and not overwrite:
        record["status"] = "exists"
        record["method"] = "skip"
        return record

    if dry_run:
        record["status"] = "downloaded"
        record["method"] = "direct" if is_direct_pdf(task["link"]) else "chrome"
        return record

    os.makedirs(policy_folder, exist_ok=True)
    # Avoid clobbering an existing differently-sourced file when re-running.
    dest_path = unique_path(dest_path)
    record["dest"] = dest_path

    method = "direct" if is_direct_pdf(task["link"]) else "chrome"
    record["method"] = method
    try:
        if method == "direct":
            asyncio.run(_download_direct_async(task["link"], dest_path))
        else:
            if driver is None:
                raise RuntimeError("Chrome driver not available")
            render_pdf_via_chrome(driver, task["link"], dest_path)
        record["status"] = "downloaded"
    except Exception as exc:  # noqa: BLE001 — record every failure, keep going
        record["status"] = "failed"
        record["error"] = f"{type(exc).__name__}: {exc}"
        # Clean up a partial/empty file so re-runs retry cleanly.
        try:
            if os.path.exists(dest_path) and os.path.getsize(dest_path) == 0:
                os.remove(dest_path)
        except OSError:
            pass
    return record


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Download a PDF for every adopted policy in a scraped workbook.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--input", "-i", required=True,
                        help="Path to the scraped .xlsx workbook")
    parser.add_argument("--sheet", "-s", required=True,
                        help="Sheet name to process (e.g. Caden)")
    parser.add_argument("--output-dir", "-o", default=".",
                        help="Where to create the 'Ten Strands' parent folder")
    parser.add_argument("--limit", type=int, default=None,
                        help="Stop after this many download tasks (test runs)")
    parser.add_argument("--start-row", type=int, default=None,
                        help="First workbook data row to process (1-based)")
    parser.add_argument("--end-row", type=int, default=None,
                        help="Last workbook data row to process (1-based)")
    parser.add_argument("--overwrite", action="store_true",
                        help="Re-download even if the target PDF already exists")
    parser.add_argument("--dry-run", "-n", action="store_true",
                        help="Plan only — create folders but do not download")
    parser.add_argument("--chrome-version", type=int, default=None,
                        help="Pin a Chrome major version (e.g. 149). Auto-detects if omitted.")
    args = parser.parse_args()

    if not os.path.isfile(args.input):
        sys.exit(f"Error: workbook not found: {args.input}")

    print(f"Reading {args.input}  [sheet: {args.sheet}]")
    try:
        tasks = load_tasks(args.input, args.sheet,
                           start_row=args.start_row, end_row=args.end_row)
    except ValueError as exc:
        sys.exit(str(exc))

    if args.limit:
        tasks = tasks[: args.limit]

    if not tasks:
        print("No adopted policies with a real link found. Nothing to do.")
        return

    districts = []
    seen = set()
    for t in tasks:
        if t["district"] not in seen:
            seen.add(t["district"])
            districts.append(t["district"])

    print(f"Found {len(tasks)} adopted policy link(s) across {len(districts)} district(s).")

    base_dir = os.path.join(args.output_dir, PARENT_FOLDER)
    os.makedirs(base_dir, exist_ok=True)
    print(f"Output folder: {os.path.abspath(base_dir)}")
    if args.dry_run:
        print("*** DRY-RUN — folders planned, nothing will be downloaded ***")

    # Decide whether we need a browser at all.
    need_chrome = any(not is_direct_pdf(t["link"]) for t in tasks) and not args.dry_run

    driver = None
    get_driver = None
    if need_chrome:
        import undetected_chromedriver as uc
        def _create_driver():
            print("Initializing Chrome for HTML->PDF rendering...")
            options = uc.ChromeOptions()
            kwargs = {"options": options}
            if args.chrome_version:
                kwargs["version_main"] = args.chrome_version
            d = uc.Chrome(**kwargs)
            d.set_page_load_timeout(45)
            return d
        get_driver = _create_driver
        driver = get_driver()

    # ── Run ──────────────────────────────────────────────────────────────────
    results = []
    counts = {"downloaded": 0, "exists": 0, "failed": 0}
    try:
        total = len(tasks)
        for i, task in enumerate(tasks, 1):
            # Preemptively recycle the browser every 100 tasks to prevent memory leaks
            if need_chrome and i > 1 and i % 100 == 0:
                print("Recycling Chrome instance to free memory...")
                try:
                    driver.quit()
                except Exception:
                    pass
                driver = get_driver()

            prefix = f"[{i}/{total}] {task['district']} | {task['policy_code']}"
            
            for attempt in range(2):
                rec = process_task(driver, task, base_dir,
                                   overwrite=args.overwrite, dry_run=args.dry_run)
                
                # If the browser crashed, restart it and retry the task once
                if rec["status"] == "failed" and "InvalidSessionIdException" in str(rec.get("error", "")):
                    if attempt == 0 and need_chrome:
                        print(f"  Browser crashed (InvalidSessionIdException). Restarting Chrome and retrying...")
                        try:
                            driver.quit()
                        except Exception:
                            pass
                        driver = get_driver()
                        continue
                break

            results.append(rec)
            counts[rec["status"]] = counts.get(rec["status"], 0) + 1

            if rec["status"] == "downloaded":
                print(f"{prefix} -> OK ({rec['method']})")
            elif rec["status"] == "exists":
                print(f"{prefix} -> exists, skipped")
            else:
                print(f"{prefix} -> FAILED: {rec['error']}")

            # Human-paced delay between requests.
            if i < total:
                time.sleep(random.uniform(_PACE_MIN, _PACE_MAX))
    finally:
        if driver is not None:
            try:
                driver.quit()
            except Exception:
                pass

    # ── Manifest + summary ───────────────────────────────────────────────────
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    manifest_path = os.path.join(base_dir, f"download_manifest_{ts}.json")
    with open(manifest_path, "w", encoding="utf-8") as fh:
        json.dump(results, fh, indent=2)
    print(f"\nManifest written: {manifest_path}")

    print(f"\n{'═' * 60}")
    print("  Summary")
    print(f"{'═' * 60}")
    print(f"  Downloaded : {counts.get('downloaded', 0)}")
    print(f"  Exists     : {counts.get('exists', 0)}")
    print(f"  Failed     : {counts.get('failed', 0)}")
    if args.dry_run:
        print("  (dry-run: no files were actually downloaded)")
    print()


if __name__ == "__main__":
    main()
"""
reader.py — Reads the tracking xlsx and parses it into DistrictRecord objects.
"""

import openpyxl
import re
from typing import Optional
from models import DistrictRecord, PolicyEntry, POLICY_DEFS

def _parse_simbli_id(url: Optional[str]) -> Optional[str]:
    if not url or "simbli" not in url.lower():
        return None
    m = re.search(r"S=(\d+)", url)
    return m.group(1) if m else None

def _parse_boarddocs_slug(url: Optional[str]) -> Optional[str]:
    if not url or "boarddocs.com" not in url.lower():
        return None
    m = re.search(r"boarddocs\.com/\w+/([^/]+)/", url)
    return m.group(1) if m else None

def load_districts(filepath: str, sheet_name: str, limit: Optional[int] = None) -> list[DistrictRecord]:
    """
    Load districts from the specified Excel file and sheet.
    Rows 1-2 are headers. Data starts at row 3.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {filepath}")
    
    ws = wb[sheet_name]
    districts = []
    
    # Iterate over data rows
    for row_idx, row_tuple in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        # Stop at empty rows or if limit reached
        cds_code = str(row_tuple[1]).strip() if row_tuple[1] is not None else ""
        if not cds_code or cds_code == "None":
            continue
            
        district_name = str(row_tuple[3]).strip() if row_tuple[3] is not None else ""
        county = str(row_tuple[2]).strip() if row_tuple[2] is not None else ""
        tracking_status = str(row_tuple[0]).strip() if row_tuple[0] is not None else None
        
        simbli_id = None
        boarddocs_slug = None
        
        policies = []
        for pdef in POLICY_DEFS:
            col_idx = pdef["col_start"] # 1-based col index of the 'value' column
            val_idx = col_idx - 1       # 0-based tuple index
            
            # Extract the quad
            val = str(row_tuple[val_idx]).strip() if row_tuple[val_idx] is not None else None
            adopted = str(row_tuple[val_idx + 1]).strip() if row_tuple[val_idx + 1] is not None else None
            revised = str(row_tuple[val_idx + 2]).strip() if row_tuple[val_idx + 2] is not None else None
            link = str(row_tuple[val_idx + 3]).strip() if row_tuple[val_idx + 3] is not None else None
            
            entry = PolicyEntry(
                col_start=col_idx,
                policy_code=pdef["code"],
                policy_title=pdef["title"],
                value=val,
                year_adopted=adopted,
                year_revised=revised,
                link=link
            )
            policies.append(entry)
            
            # Check for simbli/boarddocs IDs if not found yet
            if not simbli_id and entry.is_simbli:
                simbli_id = _parse_simbli_id(link)
            if not boarddocs_slug and entry.is_boarddocs:
                boarddocs_slug = _parse_boarddocs_slug(link)
                
        record = DistrictRecord(
            row_index=row_idx,
            cds_code=cds_code,
            county=county,
            district_name=district_name,
            tracking_status=tracking_status,
            simbli_id=simbli_id,
            boarddocs_slug=boarddocs_slug,
            policies=policies
        )
        districts.append(record)
        
        if limit and len(districts) >= limit:
            break
            
    return districts

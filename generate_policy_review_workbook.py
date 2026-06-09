"""
Generate a review workbook for the Caden tab of the Summer 2026 tracker.

The original workbook is not modified. This script checks rows 9-315 and
columns E-BH, writes proposed policy value/year/link updates into a new .xlsx,
bolds changed cells, highlights revisions/found policies green, highlights
expired links red, and adds a Summary of Changes column.

This is intentionally heuristic: district policy systems and search results vary.
Review the generated workbook before copying changes back to the source tracker.
"""

from __future__ import annotations

import argparse
import json
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from html import unescape
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl\n"
        "Install it with: python -m pip install openpyxl"
    ) from exc


DEFAULT_INPUT = Path(
    r"C:\Users\caden\Downloads\Summer 2026 Board Policy Indicator Refresh Data Tracker.xlsx"
)
DEFAULT_OUTPUT = Path("caden_policy_review_rows_9_315.xlsx")
DEFAULT_CACHE = Path("policy_review_cache.json")

SHEET_NAME = "Caden"
HEADER_ROW = 2
START_ROW = 9
END_ROW = 315
START_COL = 5  # E
END_COL = 60  # BH
GROUP_WIDTH = 4
REQUEST_DELAY_SECONDS = 1.0
TIMEOUT_SECONDS = 20

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
CHANGED_FONT = Font(bold=True)

SEARCH_URL = "https://duckduckgo.com/html/?q={query}"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125 Safari/537.36"
)


@dataclass
class PolicyGroup:
    value_col: int
    adopted_col: int
    revised_col: int
    link_col: int
    label: str
    code: str
    title: str


@dataclass
class FetchResult:
    ok: bool
    final_url: str
    status: int | None
    text: str
    error: str | None = None


def normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def is_url(value: str) -> bool:
    return value.lower().startswith(("http://", "https://"))


def extract_policy_parts(label: str) -> tuple[str, str]:
    base = re.split(r"\s+[-\u2013]\s+", label, maxsplit=1)[0].strip()
    match = re.match(r"^(BP|AR)\s+([0-9.]+)\s+(.+)$", base, re.IGNORECASE)
    if not match:
        return "", base
    return f"{match.group(1).upper()} {match.group(2)}", match.group(3).strip()


def load_cache(path: Path) -> dict:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def save_cache(path: Path, cache: dict) -> None:
    path.write_text(json.dumps(cache, indent=2, sort_keys=True), encoding="utf-8")


def request_url(url: str, timeout: int = TIMEOUT_SECONDS) -> FetchResult:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as response:
            raw = response.read(800_000)
            charset = response.headers.get_content_charset() or "utf-8"
            text = raw.decode(charset, errors="replace")
            return FetchResult(
                ok=200 <= response.status < 400,
                final_url=response.geturl(),
                status=response.status,
                text=text,
            )
    except urllib.error.HTTPError as exc:
        body = exc.read(200_000).decode("utf-8", errors="replace")
        return FetchResult(False, url, exc.code, body, str(exc))
    except Exception as exc:
        return FetchResult(False, url, None, "", str(exc))


def cached_fetch(url: str, cache: dict) -> FetchResult:
    key = f"fetch::{url}"
    if key not in cache:
        result = request_url(url)
        cache[key] = result.__dict__
        time.sleep(REQUEST_DELAY_SECONDS)
    item = cache[key]
    return FetchResult(**item)


def page_looks_expired(result: FetchResult) -> bool:
    if not result.ok:
        return True
    text = result.text.lower()
    expired_markers = [
        "404",
        "not found",
        "page cannot be found",
        "policy not found",
        "policy is no longer available",
        "expired",
        "access denied",
        "permission denied",
        "private?open",
    ]
    return any(marker in text for marker in expired_markers)


def visible_text(html: str) -> str:
    html = re.sub(r"(?is)<script.*?</script>|<style.*?</style>", " ", html)
    html = re.sub(r"(?s)<[^>]+>", " ", html)
    return re.sub(r"\s+", " ", unescape(html)).strip()


def extract_years(text: str) -> tuple[str | None, str | None]:
    adopted = None
    revised = None
    adopted_patterns = [
        r"\bAdopted\s*:?\s*(?:[^0-9]{0,30})((?:19|20)\d{2})",
        r"\bApproved\s*:?\s*(?:[^0-9]{0,30})((?:19|20)\d{2})",
    ]
    revised_patterns = [
        r"\bRevised\s*:?\s*(?:[^0-9]{0,30})((?:19|20)\d{2})",
        r"\bLast\s+Revised\s*:?\s*(?:[^0-9]{0,30})((?:19|20)\d{2})",
        r"\bUpdated\s*:?\s*(?:[^0-9]{0,30})((?:19|20)\d{2})",
    ]
    for pattern in adopted_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            adopted = match.group(1)
            break
    revised_matches: list[str] = []
    for pattern in revised_patterns:
        revised_matches.extend(re.findall(pattern, text, re.IGNORECASE))
    if revised_matches:
        revised = max(revised_matches)
    return adopted, revised


def html_search(query: str, cache: dict, max_results: int = 8) -> list[str]:
    key = f"search::{query}"
    if key in cache:
        return cache[key]

    url = SEARCH_URL.format(query=urllib.parse.quote_plus(query))
    result = request_url(url)
    links: list[str] = []
    if result.ok:
        for raw in re.findall(r'href="([^"]+)"', result.text):
            raw = unescape(raw)
            parsed = urllib.parse.urlparse(raw)
            if parsed.path == "/l/":
                params = urllib.parse.parse_qs(parsed.query)
                raw = params.get("uddg", [raw])[0]
            if raw.startswith(("http://", "https://")) and "duckduckgo.com" not in raw:
                links.append(raw)

    deduped = []
    seen = set()
    for link in links:
        clean = link.split("&rut=")[0]
        if clean not in seen:
            deduped.append(clean)
            seen.add(clean)
        if len(deduped) >= max_results:
            break

    cache[key] = deduped
    time.sleep(REQUEST_DELAY_SECONDS)
    return deduped


def search_queries(district: str, policy: PolicyGroup) -> Iterable[str]:
    quoted_district = f'"{district}"'
    yield f'{quoted_district} "{policy.code}" "{policy.title}" board policy'
    yield f'{quoted_district} "{policy.code}" board policy'
    yield f'{quoted_district} "{policy.title}" "simbli"'
    yield f'{quoted_district} "{policy.code}" site:simbli.eboardsolutions.com'
    yield f'{quoted_district} "{policy.code}" site:go.boarddocs.com'


def result_matches_policy(url: str, text: str, policy: PolicyGroup) -> bool:
    haystack = f"{url} {text}".lower()
    compact = haystack.replace(" ", "")
    code = policy.code.lower()
    code_without_space = code.replace(" ", "")
    return code in haystack or code_without_space in compact


def find_policy_link(district: str, policy: PolicyGroup, cache: dict) -> tuple[str | None, str | None, str | None]:
    for query in search_queries(district, policy):
        for candidate in html_search(query, cache):
            fetched = cached_fetch(candidate, cache)
            if page_looks_expired(fetched):
                continue
            text = visible_text(fetched.text)
            if result_matches_policy(fetched.final_url, text, policy):
                adopted, revised = extract_years(text)
                return fetched.final_url, adopted, revised
    return None, None, None


def district_policy_database_exists(district: str, cache: dict) -> bool:
    queries = [
        f'"{district}" "board policy"',
        f'"{district}" "simbli"',
        f'"{district}" "BoardDocs"',
    ]
    for query in queries:
        for link in html_search(query, cache, max_results=5):
            lowered = link.lower()
            if any(host in lowered for host in ["simbli.eboardsolutions.com", "go.boarddocs.com", "gamutonline.net"]):
                return True
    return False


def as_int_year(value: str) -> int | None:
    match = re.search(r"(19|20)\d{2}", value)
    if not match:
        return None
    return int(match.group(0))


def build_policy_groups(ws) -> list[PolicyGroup]:
    groups: list[PolicyGroup] = []
    for col in range(START_COL, END_COL + 1, GROUP_WIDTH):
        label = normalize_cell(ws.cell(HEADER_ROW, col).value)
        code, title = extract_policy_parts(label)
        groups.append(
            PolicyGroup(
                value_col=col,
                adopted_col=col + 1,
                revised_col=col + 2,
                link_col=col + 3,
                label=label,
                code=code,
                title=title,
            )
        )
    return groups


def mark_cell(cell, fill: PatternFill) -> None:
    cell.font = CHANGED_FONT
    cell.fill = fill


def set_changed(ws, row: int, col: int, value: str, fill: PatternFill) -> bool:
    cell = ws.cell(row, col)
    old = normalize_cell(cell.value)
    if old == normalize_cell(value):
        return False
    cell.value = value
    mark_cell(cell, fill)
    return True


def process_existing_policy(ws, row: int, policy: PolicyGroup, district: str, cache: dict) -> list[str]:
    notes: list[str] = []
    link = normalize_cell(ws.cell(row, policy.link_col).value)
    old_adopted = normalize_cell(ws.cell(row, policy.adopted_col).value)
    old_revised = normalize_cell(ws.cell(row, policy.revised_col).value)

    if not is_url(link):
        found_link, found_adopted, found_revised = find_policy_link(district, policy, cache)
        if found_link:
            set_changed(ws, row, policy.link_col, found_link, GREEN_FILL)
            if found_adopted:
                set_changed(ws, row, policy.adopted_col, found_adopted, GREEN_FILL)
            if found_revised:
                set_changed(ws, row, policy.revised_col, found_revised, GREEN_FILL)
            notes.append(f"{policy.code}: found replacement link for existing policy")
        return notes

    fetched = cached_fetch(link, cache)
    if page_looks_expired(fetched):
        mark_cell(ws.cell(row, policy.value_col), RED_FILL)
        mark_cell(ws.cell(row, policy.link_col), RED_FILL)
        found_link, found_adopted, found_revised = find_policy_link(district, policy, cache)
        if found_link:
            set_changed(ws, row, policy.link_col, found_link, RED_FILL)
            if found_adopted:
                set_changed(ws, row, policy.adopted_col, found_adopted, RED_FILL)
            if found_revised:
                set_changed(ws, row, policy.revised_col, found_revised, RED_FILL)
            notes.append(f"{policy.code}: original link expired; replacement found")
        else:
            replacement = "N/A" if district_policy_database_exists(district, cache) else "*"
            set_changed(ws, row, policy.link_col, replacement, RED_FILL)
            notes.append(f"{policy.code}: original link expired; no replacement found")
        return notes

    text = visible_text(fetched.text)
    adopted, revised = extract_years(text)
    changed_parts: list[str] = []
    if adopted and old_adopted in {"", "N/A", "*"}:
        if set_changed(ws, row, policy.adopted_col, adopted, GREEN_FILL):
            changed_parts.append(f"adopted year {old_adopted or 'blank'} -> {adopted}")
    if revised:
        old_year = as_int_year(old_revised)
        new_year = as_int_year(revised)
        if new_year and (old_year is None or new_year > old_year):
            if set_changed(ws, row, policy.revised_col, revised, GREEN_FILL):
                changed_parts.append(f"revised year {old_revised or 'blank'} -> {revised}")
    final_url = fetched.final_url
    if final_url and final_url != link:
        if set_changed(ws, row, policy.link_col, final_url, GREEN_FILL):
            changed_parts.append("canonical link updated")

    if changed_parts:
        mark_cell(ws.cell(row, policy.value_col), GREEN_FILL)
        notes.append(f"{policy.code}: " + "; ".join(changed_parts))
    return notes


def process_missing_policy(ws, row: int, policy: PolicyGroup, district: str, cache: dict) -> list[str]:
    found_link, adopted, revised = find_policy_link(district, policy, cache)
    if not found_link:
        return []

    set_changed(ws, row, policy.value_col, "1", GREEN_FILL)
    set_changed(ws, row, policy.link_col, found_link, GREEN_FILL)
    if adopted:
        set_changed(ws, row, policy.adopted_col, adopted, GREEN_FILL)
    if revised:
        set_changed(ws, row, policy.revised_col, revised, GREEN_FILL)
    return [f"{policy.code}: policy found and proposed as passed"]


def generate_review(
    input_path: Path,
    output_path: Path,
    cache_path: Path,
    start_row: int,
    end_row: int,
) -> None:
    wb = load_workbook(input_path)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {SHEET_NAME}")
    ws = wb[SHEET_NAME]
    groups = build_policy_groups(ws)
    cache = load_cache(cache_path)

    summary_col = END_COL + 1
    ws.cell(HEADER_ROW, summary_col).value = "Summary of Changes"
    ws.cell(HEADER_ROW, summary_col).font = Font(bold=True)
    ws.column_dimensions[get_column_letter(summary_col)].width = 80

    for row in range(start_row, end_row + 1):
        district = normalize_cell(ws.cell(row, 4).value)
        if not district:
            continue

        row_notes: list[str] = []
        print(f"Row {row}: {district}")
        for policy in groups:
            value = normalize_cell(ws.cell(row, policy.value_col).value)
            if value in {"1", "1.0"}:
                row_notes.extend(process_existing_policy(ws, row, policy, district, cache))
            else:
                row_notes.extend(process_missing_policy(ws, row, policy, district, cache))
            save_cache(cache_path, cache)

        if row_notes:
            summary = " | ".join(row_notes)
            ws.cell(row, summary_col).value = summary
            mark_cell(ws.cell(row, summary_col), YELLOW_FILL)

        wb.save(output_path)

    wb.save(output_path)


def main() -> None:
    global REQUEST_DELAY_SECONDS

    parser = argparse.ArgumentParser(description="Generate Caden policy review workbook.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--cache", type=Path, default=DEFAULT_CACHE)
    parser.add_argument("--start-row", type=int, default=START_ROW)
    parser.add_argument("--end-row", type=int, default=END_ROW)
    parser.add_argument("--delay", type=float, default=REQUEST_DELAY_SECONDS)
    args = parser.parse_args()

    REQUEST_DELAY_SECONDS = args.delay

    generate_review(args.input, args.output, args.cache, args.start_row, args.end_row)
    print(f"Review workbook written to: {args.output.resolve()}")


if __name__ == "__main__":
    main()

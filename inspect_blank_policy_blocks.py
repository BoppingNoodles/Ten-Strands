from openpyxl import load_workbook


WORKBOOK = "Summer_2026_Scraped_20260630_211511_blanks_fixed.xlsx"
SHEET = "Caden"
START_ROW = 121
END_ROW = 200
POLICY_COLS = [
    (5, "BP3510"),
    (9, "BP3511"),
    (13, "BP3511.1"),
    (17, "BP3514"),
    (21, "BP3514.1"),
    (25, "BP5142.2"),
    (29, "BP6142.5"),
    (33, "BP7110"),
    (37, "AR3511.1"),
    (41, "AR3514"),
    (45, "AR3514.1"),
    (49, "AR3514.2"),
    (53, "AR5142.2"),
    (57, "AR7110"),
]


def is_blank(value):
    return value is None or str(value).strip() == ""


wb = load_workbook(WORKBOOK, read_only=True, data_only=False)
ws = wb[SHEET]
blanks = []

for row_idx, row in enumerate(
    ws.iter_rows(min_row=START_ROW, max_row=END_ROW, values_only=True),
    start=START_ROW,
):
    district = row[3]
    for col_start, policy_name in POLICY_COLS:
        values = row[col_start - 1 : col_start + 3]
        if all(is_blank(value) for value in values):
            blanks.append((row_idx, district, policy_name, col_start))

print(f"blank_blocks={len(blanks)}")
for item in blanks:
    print(item)

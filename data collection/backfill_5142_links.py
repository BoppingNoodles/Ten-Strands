"""
backfill_5142_links.py — Targeted backfill for BP/AR 5142.2 rows that have
value=1 but a missing/empty link in the Data Tracker.

Reads the tracker, identifies all rows where BP 5142.2 (col 25) or
AR 5142.2 (col 53) has value=1 but no http link, then runs a Simbli
policy-index lookup for each and writes the result back.

Only modifies the specified --sheet (default: Caden). Other sheets are
loaded unchanged and saved back intact.

Usage:
    python backfill_5142_links.py [--input PATH] [--output PATH] [--sheet SHEET]
"""

import argparse
import re
import random
import time

import openpyxl
import undetected_chromedriver as uc

import simbli as _simbli
from models import POLICY_DEFS

# ── Constants ─────────────────────────────────────────────────────────────────

BP_5142_DEF = next(p for p in POLICY_DEFS if p["code"] == "BP 5142.2")
AR_5142_DEF = next(p for p in POLICY_DEFS if p["code"] == "AR 5142.2")

SIMBLI_HOME = "https://simbli.eboardsolutions.com/"

# ── Helpers ───────────────────────────────────────────────────────────────────

def _s(v) -> str:
    return str(v).strip() if v is not None else ""


def _parse_simbli_id(url: str | None) -> str | None:
    if not url or "simbli" not in url.lower():
        return None
    m = re.search(r"S=(\d+)", url)
    return m.group(1) if m else None


def _needs_link(val, link_val) -> bool:
    """Return True if the policy is adopted (val=1) but lacks an http link."""
    v = _s(val)
    return v in ("1", "1.0") and not _s(link_val).startswith("http")


# ── Self-healing browser session ──────────────────────────────────────────────

class BrowserSession:
    def __init__(self):
        self.driver = self._create_driver()

    def _create_driver(self):
        print("  [Browser] Initializing Chrome...")
        options = uc.ChromeOptions()
        d = uc.Chrome(options=options, version_main=149)
        d.set_page_load_timeout(45)
        return d

    def ensure_alive(self):
        try:
            _ = self.driver.current_url
            return self.driver
        except Exception:
            print("  [Browser] Chrome session died. Restarting...")
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = self._create_driver()
            # Re-warm the new session
            _warm_up(self.driver)
            return self.driver

    def quit(self):
        try:
            self.driver.quit()
        except Exception:
            pass


# ── Simbli session warm-up ────────────────────────────────────────────────────

def _warm_up(driver) -> None:
    """
    Navigate to the Simbli home page so the session gets the necessary cookies
    (CoreAuthToken etc.) before making API calls against individual district
    policy indexes. Without this, the PolicyListing API returns empty.
    """
    print("  [Browser] Warming up Simbli session...")
    try:
        driver.get(SIMBLI_HOME)
        _simbli._wait_for_page(driver, extra_wait=3.0)
        print("  [Browser] Warm-up done.")
    except Exception as e:
        print(f"  [Browser] Warm-up failed (non-fatal): {e}")


# ── Row loader ────────────────────────────────────────────────────────────────

def load_problem_rows(ws) -> list[dict]:
    """
    Scan the sheet and return a list of dicts for every row where BP or AR
    5142.2 is adopted but missing a link.
    """
    problem_rows = []

    for row_idx in range(3, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, 80)]
        cds = _s(row[1])
        if not cds or cds == "None":
            continue

        district_name = _s(row[3])

        # BP 5142.2: col_start=25 → 0-based idx 24; link offset+3 → idx 27
        bp_val  = row[24]
        bp_link = row[27]
        # AR 5142.2: col_start=53 → 0-based idx 52; link offset+3 → idx 55
        ar_val  = row[52]
        ar_link = row[55]

        bp_needs = _needs_link(bp_val, bp_link)
        ar_needs = _needs_link(ar_val, ar_link)

        if not bp_needs and not ar_needs:
            continue

        # Infer simbli_id from any existing http link in the row
        simbli_id = None
        for pdef in POLICY_DEFS:
            col_start = pdef["col_start"]
            link_cell = row[col_start + 3 - 1]  # 0-based: (col_start+3) - 1
            sid = _parse_simbli_id(_s(link_cell))
            if sid:
                simbli_id = sid
                break

        problem_rows.append({
            "row_idx":       row_idx,
            "cds_code":      cds,
            "district_name": district_name,
            "simbli_id":     simbli_id,
            "bp_needs_link": bp_needs,
            "ar_needs_link": ar_needs,
            "bp_col_start":  BP_5142_DEF["col_start"],   # 25
            "ar_col_start":  AR_5142_DEF["col_start"],   # 53
        })

    return problem_rows


# ── Simbli lookup ─────────────────────────────────────────────────────────────

def _lookup_5142_links(simbli_id: str, district_name: str, browser: BrowserSession) -> dict:
    """
    Fetch the Simbli policy index for `simbli_id` and return:
        { "BP 5142.2": <url|None>, "AR 5142.2": <url|None> }

    Uses the self-healing BrowserSession; clears the index cache entry
    first so we get a fresh fetch (not stale None from a previous crash).
    """
    # Clear any cached empty result from a previous failed attempt
    _simbli._INDEX_CACHE.pop(simbli_id, None)

    driver = browser.ensure_alive()
    rows = _simbli._get_policy_listing_sync(driver, simbli_id)

    if not rows:
        print(f"    [WARN] Empty index for {district_name} (simbli_id={simbli_id})")
        return {"BP 5142.2": None, "AR 5142.2": None}

    bp_match = _simbli._find_matching_policy(rows, "BP 5142.2")
    ar_match = _simbli._find_matching_policy(rows, "AR 5142.2")

    bp_link = bp_match.get("link") if bp_match else None
    ar_link = ar_match.get("link") if ar_match else None

    print(f"    BP 5142.2 -> {bp_link!r}")
    print(f"    AR 5142.2 -> {ar_link!r}")
    return {"BP 5142.2": bp_link, "AR 5142.2": ar_link}


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Backfill missing BP/AR 5142.2 links")
    parser.add_argument(
        "--input",
        default="Summer 2026 Board Policy Indicator Refresh Data Tracker.xlsx",
        help="Source workbook",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output path (default: overwrites --input)",
    )
    parser.add_argument(
        "--sheet",
        default="Caden",
        help="Sheet name to update (only this sheet is modified)",
    )
    args = parser.parse_args()

    output_path = args.output or args.input

    print(f"Loading '{args.input}' sheet '{args.sheet}'...")
    wb = openpyxl.load_workbook(args.input)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet '{args.sheet}' not found. Available: {wb.sheetnames}")
    ws = wb[args.sheet]

    problem_rows = load_problem_rows(ws)
    if not problem_rows:
        print("No rows need BP/AR 5142.2 link backfill. Done.")
        return

    print(f"\nFound {len(problem_rows)} rows needing backfill:")
    for pr in problem_rows:
        flags = []
        if pr["bp_needs_link"]:
            flags.append("BP 5142.2")
        if pr["ar_needs_link"]:
            flags.append("AR 5142.2")
        sid = pr["simbli_id"] or "NO SIMBLI ID"
        print(f"  Row {pr['row_idx']:>3}: {pr['district_name']} ({sid}) -> {', '.join(flags)}")

    # Group by simbli_id so we only fetch each index once
    by_simbli: dict[str, list[dict]] = {}
    no_simbli = []
    for pr in problem_rows:
        sid = pr["simbli_id"]
        if sid:
            by_simbli.setdefault(sid, []).append(pr)
        else:
            no_simbli.append(pr)

    if no_simbli:
        print(f"\n[WARN] {len(no_simbli)} row(s) have no simbli_id and will be skipped:")
        for pr in no_simbli:
            print(f"  Row {pr['row_idx']:>3}: {pr['district_name']}")

    print(f"\nStarting browser ({len(by_simbli)} unique Simbli IDs to query)...")
    browser = BrowserSession()
    _warm_up(browser.driver)

    updated = []
    skipped = []

    try:
        total = len(by_simbli)
        for i, (simbli_id, rows_for_sid) in enumerate(by_simbli.items(), 1):
            district_name = rows_for_sid[0]["district_name"]
            print(f"\n[{i}/{total}] [{simbli_id}] {district_name}...")

            links = _lookup_5142_links(simbli_id, district_name, browser)

            for pr in rows_for_sid:
                row_idx = pr["row_idx"]
                changed = False

                if pr["bp_needs_link"]:
                    new_link = links["BP 5142.2"]
                    if new_link and new_link.startswith("http"):
                        ws.cell(row=row_idx, column=pr["bp_col_start"] + 3).value = new_link
                        print(f"    Row {row_idx} BP 5142.2 link written.")
                        changed = True
                    else:
                        print(f"    Row {row_idx} BP 5142.2: no link found in index.")
                        skipped.append((row_idx, pr["district_name"], "BP 5142.2"))

                if pr["ar_needs_link"]:
                    new_link = links["AR 5142.2"]
                    if new_link and new_link.startswith("http"):
                        ws.cell(row=row_idx, column=pr["ar_col_start"] + 3).value = new_link
                        print(f"    Row {row_idx} AR 5142.2 link written.")
                        changed = True
                    else:
                        print(f"    Row {row_idx} AR 5142.2: no link found in index.")
                        skipped.append((row_idx, pr["district_name"], "AR 5142.2"))

                if changed:
                    updated.append((row_idx, pr["district_name"]))

            # Polite delay between districts to avoid bot detection
            if i < total:
                pause = random.uniform(2.5, 5.0)
                print(f"  [pause {pause:.1f}s]")
                time.sleep(pause)

    finally:
        print("\nClosing browser...")
        browser.quit()

    print(f"\nSaving to '{output_path}'...")
    wb.save(output_path)

    print(f"\n=== Backfill complete ===")
    print(f"  Sheet modified            : {args.sheet} only")
    print(f"  Rows with link(s) written : {len(set(r for r, _ in updated))}")
    print(f"  Policies still missing    : {len(skipped)}")
    if skipped:
        print("  Still missing:")
        for row_idx, dist, code in skipped:
            print(f"    Row {row_idx}: {dist} / {code}")


if __name__ == "__main__":
    main()

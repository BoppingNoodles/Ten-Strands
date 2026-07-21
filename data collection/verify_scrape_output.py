import json
from pathlib import Path

from openpyxl import load_workbook


LOG_PATH = Path("scrape_log_20260630_155709.json")
WORKBOOK_PATH = Path("Summer_2026_Scraped_20260630_155709.xlsx")
SHEET_NAME = "Caden"

GREEN_RGBS = {"0000B050", "FF00B050"}
RED_RGBS = {"00FF0000", "FFFF0000"}


def cell_rgb(cell):
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return None
    return fill.fgColor.rgb


def main() -> None:
    data = json.loads(LOG_PATH.read_text(encoding="utf-8"))
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]

    row_by_cds = {}
    for row in range(3, ws.max_row + 1):
        cds = ws.cell(row, 2).value
        if cds:
            row_by_cds[str(cds).strip()] = row

    districts = []
    seen = set()
    mismatches = []
    highlighted_expected = 0
    highlighted_ok = 0
    changes_expected = 0
    changes_ok = 0

    for entry in data:
        key = (entry["cds_code"], entry["district_name"])
        if key not in seen:
            seen.add(key)
            districts.append(key)

        row = row_by_cds.get(str(entry["cds_code"]).strip())
        if not row:
            mismatches.append((entry["cds_code"], entry["policy_code"], "missing row"))
            continue

        col = int(entry["col_start"])
        for field, offset in [
            ("new_value", 0),
            ("new_year_revised", 2),
            ("new_link", 3),
        ]:
            new_value = entry.get(field)
            if new_value is None:
                continue
            changes_expected += 1
            actual = ws.cell(row, col + offset).value
            if str(actual).strip() == str(new_value).strip():
                changes_ok += 1
            else:
                mismatches.append((row, entry["policy_code"], field, new_value, actual))

        color = entry.get("highlight_color")
        if color in {"green", "red"}:
            highlighted_expected += 1
            rgb = cell_rgb(ws.cell(row, col))
            if color == "green" and rgb in GREEN_RGBS:
                highlighted_ok += 1
            elif color == "red" and rgb in RED_RGBS:
                highlighted_ok += 1
            else:
                mismatches.append((row, entry["policy_code"], "highlight", color, rgb))

    first_row = row_by_cds.get(str(districts[0][0]).strip())
    last_row = row_by_cds.get(str(districts[-1][0]).strip())

    print(f"log_entries={len(data)}")
    print(f"district_count={len(districts)}")
    print(f"first_district={districts[0]} row={first_row}")
    print(f"last_district={districts[-1]} row={last_row}")
    print(f"changes_checked={changes_ok}/{changes_expected}")
    print(f"highlights_checked={highlighted_ok}/{highlighted_expected}")
    print(f"mismatches={len(mismatches)}")
    for mismatch in mismatches[:25]:
        print("mismatch:", mismatch)


if __name__ == "__main__":
    main()

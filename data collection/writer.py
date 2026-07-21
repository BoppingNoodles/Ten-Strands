"""
writer.py — Applies ScrapeResult objects back to a copy of the xlsx and saves it.
"""

import openpyxl
from openpyxl.styles import PatternFill
import json
from datetime import datetime
from models import ScrapeResult, HighlightColor, is_safe_routes_policy_code

# Highlight Colors
GREEN_FILL = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)

def write_output(input_path: str, results: list[list[ScrapeResult]], output_path: str, sheet_name: str):
    """
    Updates the workbook with the scraping results and saves to output_path.
    `results` is a list of lists (one list of ScrapeResults per district).
    """
    wb = openpyxl.load_workbook(input_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found.")
    ws = wb[sheet_name]
    
    # Map results to their row/col positions. 
    # To do this safely, we map by CDS code and column start index.
    
    # Build a row lookup by CDS code
    row_map = {}
    for row_idx, row_tuple in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        cds = str(row_tuple[1]).strip() if row_tuple[1] is not None else ""
        if cds:
            row_map[cds] = row_idx
            
    # Apply results
    for district_results in results:
        # Ignore exceptions or empty lists from gather
        if isinstance(district_results, Exception) or not district_results:
            continue
            
        cds = district_results[0].cds_code
        if cds not in row_map:
            continue
            
        row_idx = row_map[cds]
        district_updated = False
        
        for res in district_results:
            safe_routes = is_safe_routes_policy_code(res.policy_code)

            # Update cells if there are changes
            if res.new_value is not None:
                ws.cell(row=row_idx, column=res.col_start).value = res.new_value
            if res.new_year_adopted is not None:
                ws.cell(row=row_idx, column=res.col_start + 1).value = res.new_year_adopted
            if res.new_year_revised is not None:
                ws.cell(row=row_idx, column=res.col_start + 2).value = res.new_year_revised
            if res.new_link is not None:
                ws.cell(row=row_idx, column=res.col_start + 3).value = res.new_link

            if safe_routes:
                # Safe Routes updates are written but never highlighted or tracked.
                for offset in range(4):
                    ws.cell(row=row_idx, column=res.col_start + offset).fill = NO_FILL
                continue

            # Apply highlighting to the VALUE cell (col_start)
            if res.highlight_color == HighlightColor.GREEN:
                ws.cell(row=row_idx, column=res.col_start).fill = GREEN_FILL
                district_updated = True
            elif res.highlight_color == HighlightColor.RED:
                ws.cell(row=row_idx, column=res.col_start).fill = RED_FILL
                district_updated = True
            else:
                ws.cell(row=row_idx, column=res.col_start).fill = NO_FILL

            # Decide whether this result represents a real policy-status change
            # worth flagging the district as "Policy Updated". By default that
            # means an actionable highlight: revised/newly found (GREEN) or a
            # dead link (RED). Other writes (link backfill, redirect, cosmetic
            # blank normalization to 0/N/A) update the cells but are bookkeeping,
            # not a policy update. A result may override this via tracking_change.
            if res.tracking_change is True:
                district_updated = True
            elif res.tracking_change is False:
                continue
            elif res.highlight_color in (HighlightColor.GREEN, HighlightColor.RED):
                district_updated = True
                
        if district_updated:
            ws.cell(row=row_idx, column=1).value = "Policy Updated"
            
    wb.save(output_path)
    
def write_log(results: list[list[ScrapeResult]], log_path: str):
    """
    Write all scrape results to a JSON log file.
    """
    flat_results = []
    for d_results in results:
        if isinstance(d_results, Exception):
            flat_results.append({"error": str(d_results)})
        elif isinstance(d_results, list):
            for r in d_results:
                if isinstance(r, ScrapeResult):
                    flat_results.append(r.to_dict())
                
    with open(log_path, 'w', encoding='utf-8') as f:
        json.dump(flat_results, f, indent=2)

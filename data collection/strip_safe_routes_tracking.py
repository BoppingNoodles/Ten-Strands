"""Remove Safe Routes (BP/AR 5142.2) highlights and spurious Policy Updated tags."""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from models import POLICY_DEFS, SAFE_ROUTES_COL_STARTS, is_safe_routes_policy_code


WORKBOOK = "Summer_2026_Scraped_20260701_084846_blanks_fixed.xlsx"
OUTPUT = WORKBOOK
BASELINE = "Summer 2026 Board Policy Indicator Refresh Data Tracker.xlsx"
SHEET = "Caden"
START_ROW = 121
END_ROW = 200
NO_FILL = PatternFill(fill_type=None)


def cell_values_equal(left, right) -> bool:
    if left is None and right is None:
        return True
    if left is None or right is None:
        left_text = "" if left is None else str(left).strip()
        right_text = "" if right is None else str(right).strip()
        return left_text == right_text
    return str(left).strip() == str(right).strip()


def main():
    wb = load_workbook(WORKBOOK)
    baseline_wb = load_workbook(BASELINE, data_only=True)
    ws = wb[SHEET]
    baseline_ws = baseline_wb[SHEET]

    non_safe_cols = [
        p["col_start"] + offset
        for p in POLICY_DEFS
        if not is_safe_routes_policy_code(p["code"])
        for offset in range(4)
    ]
    safe_cols = [
        col_start + offset
        for col_start in SAFE_ROUTES_COL_STARTS
        for offset in range(4)
    ]

    cleared_highlights = 0
    cleared_status = []

    for row in range(START_ROW, END_ROW + 1):
        for col in safe_cols:
            cell = ws.cell(row=row, column=col)
            if cell.fill and cell.fill.fill_type:
                cell.fill = NO_FILL
                cleared_highlights += 1

        if ws.cell(row=row, column=1).value != "Policy Updated":
            continue

        non_safe_changed = any(
            not cell_values_equal(
                ws.cell(row=row, column=col).value,
                baseline_ws.cell(row=row, column=col).value,
            )
            for col in non_safe_cols
        )
        if not non_safe_changed:
            ws.cell(row=row, column=1).value = baseline_ws.cell(row=row, column=1).value
            cleared_status.append((row, ws.cell(row=row, column=4).value))

    wb.save(OUTPUT)
    print(f"Updated {OUTPUT}")
    print(f"Cleared Safe Routes cell fills: {cleared_highlights}")
    print(f"Cleared Policy Updated on {len(cleared_status)} rows (Safe Routes-only changes).")
    for row, district in cleared_status:
        print(f"  Row {row}: {district}")


if __name__ == "__main__":
    main()

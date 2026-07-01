from openpyxl import load_workbook

from models import POLICY_DEFS, is_bad_link


WORKBOOK = "Summer_2026_Scraped_20260701_084846_blanks_fixed.xlsx"
SHEET = "Caden"
START_ROW = 121
END_ROW = 200
POLICY_COLS = [(p["col_start"], p["code"]) for p in POLICY_DEFS if p["col_start"] <= 57]


def is_blank(value):
    return value is None or str(value).strip() == ""


wb = load_workbook(WORKBOOK, read_only=True, data_only=False)
ws = wb[SHEET]
blank_blocks = []
bad_links = []

for row_idx, row in enumerate(
    ws.iter_rows(min_row=START_ROW, max_row=END_ROW, values_only=True),
    start=START_ROW,
):
    district = row[3]
    for col_start, policy_name in POLICY_COLS:
        values = row[col_start - 1 : col_start + 3]
        if all(is_blank(value) for value in values):
            blank_blocks.append((row_idx, district, policy_name, col_start))
        link_value = values[3]
        if is_bad_link(link_value):
            bad_links.append((row_idx, district, policy_name, col_start + 3, link_value))

print(f"blank_blocks={len(blank_blocks)}")
for item in blank_blocks:
    print("BLANK", item)

print(f"bad_links={len(bad_links)}")
for item in bad_links:
    print("BAD_LINK", item)

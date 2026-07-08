from openpyxl import load_workbook


INPUT = "Summer_2026_Scraped_20260630_173257.xlsx"
OUTPUT = "Summer_2026_Scraped_20260630_173257_safe_routes_fixed.xlsx"
SHEET = "Caden"
START_ROW = 121
END_ROW = 200
SAFE_ROUTES_COLS = (25, 53)  # BP 5142.2, AR 5142.2 value columns


def is_blank(value):
    return value is None or str(value).strip() == ""


def main():
    wb = load_workbook(INPUT)
    ws = wb[SHEET]
    updated = []

    for row in range(START_ROW, END_ROW + 1):
        district = ws.cell(row=row, column=4).value
        for col_start in SAFE_ROUTES_COLS:
            cells = [ws.cell(row=row, column=col_start + offset) for offset in range(4)]
            if all(is_blank(cell.value) for cell in cells):
                cells[0].value = "0"
                cells[1].value = "N/A"
                cells[2].value = "N/A"
                cells[3].value = "N/A"
                updated.append((row, district, ws.cell(row=2, column=col_start).value))

    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")
    print(f"Normalized {len(updated)} blank Safe Routes policy blocks.")
    for row, district, policy in updated:
        print(f"Row {row}: {district} - {policy}")


if __name__ == "__main__":
    main()

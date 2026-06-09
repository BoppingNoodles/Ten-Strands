import json
from pathlib import Path

from openpyxl import load_workbook


cache = json.loads(Path("policy_review_cache.json").read_text(encoding="utf-8"))
keys = "\n".join(cache.keys())
wb = load_workbook(
    r"C:\Users\caden\Downloads\Summer 2026 Board Policy Indicator Refresh Data Tracker.xlsx",
    read_only=True,
)
ws = wb["Caden"]
processed = []
for row in range(9, 316):
    district = ws.cell(row, 4).value or ""
    if district and district in keys:
        processed.append((row, district))

print("processed_count", len(processed))
print("last_rows", processed[-15:])

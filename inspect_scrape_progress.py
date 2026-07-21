import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


LOG_PATH = Path("scrape_log_20260630_155709.json")
WORKBOOK_PATH = Path("Summer 2026 Board Policy Indicator Refresh Data Tracker.xlsx")
SHEET_NAME = "Caden"


def main() -> None:
    data = json.loads(LOG_PATH.read_text(encoding="utf-8"))
    wb = load_workbook(WORKBOOK_PATH, read_only=True)
    ws = wb[SHEET_NAME]

    row_by_cds = {}
    for row_cells in ws.iter_rows():
        row = row_cells[0].row
        cds = row_cells[1].value if len(row_cells) > 1 else None
        if cds is not None:
            row_by_cds[str(cds).strip()] = row

    districts = []
    seen = set()
    for entry in data:
        key = (entry["cds_code"], entry["district_name"])
        if key not in seen:
            seen.add(key)
            districts.append(key)

    counts = Counter((entry["cds_code"], entry["district_name"]) for entry in data)
    last_cds, last_name = districts[-1]
    last_row = row_by_cds.get(str(last_cds).strip())

    print(f"entries={len(data)}")
    print(f"district_count={len(districts)}")
    print(f"first={districts[0]} row={row_by_cds.get(str(districts[0][0]).strip())}")
    print(f"last={districts[-1]} row={last_row} entries_for_last={counts[(last_cds, last_name)]}")
    print("last_10:")
    for cds, name in districts[-10:]:
        print(f"  row={row_by_cds.get(str(cds).strip())} cds={cds} district={name} entries={counts[(cds, name)]}")
    print(f"resume_start_row={last_row + 1 if last_row else 'unknown'}")


if __name__ == "__main__":
    main()

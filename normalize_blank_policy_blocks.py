from openpyxl import load_workbook

from models import POLICY_DEFS, is_bad_link


INPUT = "Summer_2026_Scraped_20260701_111447.xlsx"
OUTPUT = "Summer_2026_Scraped_20260701_084846_blanks_fixed.xlsx"
SHEET = "Caden"
START_ROW = 121
END_ROW = 200


def is_blank(value):
    return value is None or str(value).strip() == ""


def main():
    wb = load_workbook(INPUT)
    ws = wb[SHEET]
    blank_updates = []
    link_updates = []

    for row in range(START_ROW, END_ROW + 1):
        district = ws.cell(row=row, column=4).value
        for pdef in POLICY_DEFS:
            if pdef["col_start"] > 57:
                continue
            col_start = pdef["col_start"]
            policy_name = pdef["code"]
            cells = [ws.cell(row=row, column=col_start + offset) for offset in range(4)]

            if all(is_blank(cell.value) for cell in cells):
                cells[0].value = "0"
                cells[1].value = "N/A"
                cells[2].value = "N/A"
                cells[3].value = "N/A"
                blank_updates.append((row, district, policy_name))
                continue

            link_cell = cells[3]
            if is_bad_link(link_cell.value):
                link_cell.value = "N/A"
                link_updates.append((row, district, policy_name, "N/A"))

    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")
    print(f"Normalized {len(blank_updates)} blank policy blocks.")
    for row, district, policy_name in blank_updates:
        print(f"BLANK Row {row}: {district} - {policy_name}")
    print(f"Cleared {len(link_updates)} invalid link cells.")
    for row, district, policy_name, new_link in link_updates:
        print(f"LINK Row {row}: {district} - {policy_name} -> {new_link}")


if __name__ == "__main__":
    main()
